"""
Weekly Project Status Report Generator
Input : data/tasks.csv, data/budget.csv, data/team.csv
Output: output/weekly_report_YYYY-MM-DD.xlsx (4 formatted sheets + charts)

Usage:
    python src/report_generator.py
"""

import time
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.styles import (Alignment, Border, Font, GradientFill,
                               PatternFill, Side)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

# ── Paths ────────────────────────────────────────────────────────────────────
BASE = Path(__file__).parent.parent
DATA = BASE / "data"
OUTPUT = BASE / "output"

# ── Colour palette ────────────────────────────────────────────────────────────
C_HEADER_BG   = "1F4E79"   # dark blue
C_HEADER_FONT = "FFFFFF"
C_ROW_ALT     = "EBF3FB"   # light blue stripe
C_DONE        = "C6EFCE"   # green fill
C_IN_PROGRESS = "FFEB9C"   # yellow fill
C_OVERDUE     = "FFC7CE"   # red fill
C_NOT_STARTED = "EDEDED"   # grey fill
C_KPI_BG      = "2E75B6"   # KPI card blue
C_KPI_WARN    = "C00000"   # KPI card red (overdue)
C_KPI_OK      = "375623"   # KPI card green

STATUS_FILLS = {
    "Done":        PatternFill("solid", fgColor=C_DONE),
    "In Progress": PatternFill("solid", fgColor=C_IN_PROGRESS),
    "Overdue":     PatternFill("solid", fgColor=C_OVERDUE),
    "Not Started": PatternFill("solid", fgColor=C_NOT_STARTED),
}

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ── Helpers ───────────────────────────────────────────────────────────────────

def header_style(cell, bg: str = C_HEADER_BG) -> None:
    cell.font = Font(bold=True, color=C_HEADER_FONT, size=11)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER


def data_style(cell, row_idx: int, fill: PatternFill | None = None) -> None:
    cell.fill = fill if fill else (
        PatternFill("solid", fgColor=C_ROW_ALT) if row_idx % 2 == 0
        else PatternFill("solid", fgColor="FFFFFF")
    )
    cell.alignment = Alignment(vertical="center")
    cell.border = BORDER


def set_col_widths(ws, widths: dict[str, int]) -> None:
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def write_df_to_sheet(ws, df: pd.DataFrame, col_widths: dict) -> None:
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                header_style(cell)
            else:
                data_style(cell, r_idx)
    ws.row_dimensions[1].height = 30
    set_col_widths(ws, col_widths)
    ws.freeze_panes = "A2"


# ── Sheet 1: Summary ──────────────────────────────────────────────────────────

def build_summary(wb: Workbook, tasks: pd.DataFrame, budget: pd.DataFrame,
                  team: pd.DataFrame) -> None:
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False

    # KPIs
    total = len(tasks)
    done = (tasks["Status"] == "Done").sum()
    overdue = (tasks["Status"] == "Overdue").sum()
    budget_pct = budget["Spent ($)"].sum() / budget["Allocated ($)"].sum() * 100
    team_util = (team["Completed"] + team["In Progress"]).sum() / team["Total Tasks"].sum() * 100

    kpis = [
        ("Tasks Completed", f"{done}/{total}  ({done/total*100:.0f}%)", C_KPI_OK),
        ("Budget Used", f"${budget['Spent ($)'].sum():,.0f} / ${budget['Allocated ($)'].sum():,.0f}  ({budget_pct:.0f}%)",
         C_KPI_WARN if budget_pct > 85 else C_KPI_BG),
        ("Overdue Tasks", str(overdue), C_KPI_WARN if overdue > 0 else C_KPI_OK),
        ("Team Utilization", f"{team_util:.0f}%", C_KPI_BG),
    ]

    ws.merge_cells("A1:F1")
    title = ws["A1"]
    title.value = "📊  Weekly Project Status Report"
    title.font = Font(bold=True, size=16, color=C_HEADER_BG)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:F2")
    ws["A2"].value = f"Generated: {date.today().strftime('%B %d, %Y')}"
    ws["A2"].font = Font(italic=True, color="808080")
    ws["A2"].alignment = Alignment(horizontal="center")

    for i, (label, value, color) in enumerate(kpis):
        col_start = i * 2 + 1
        col_end = col_start + 1
        row_label, row_value = 4, 5

        ws.merge_cells(start_row=row_label, start_column=col_start,
                       end_row=row_label, end_column=col_end)
        ws.merge_cells(start_row=row_value, start_column=col_start,
                       end_row=row_value, end_column=col_end)

        lbl_cell = ws.cell(row=row_label, column=col_start, value=label)
        lbl_cell.fill = PatternFill("solid", fgColor=color)
        lbl_cell.font = Font(bold=True, color="FFFFFF", size=10)
        lbl_cell.alignment = Alignment(horizontal="center", vertical="center")

        val_cell = ws.cell(row=row_value, column=col_start, value=value)
        val_cell.fill = PatternFill("solid", fgColor=color)
        val_cell.font = Font(bold=True, color="FFFFFF", size=13)
        val_cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[row_label].height = 22
        ws.row_dimensions[row_value].height = 32

    # Status breakdown table
    ws["A7"].value = "Status Breakdown"
    ws["A7"].font = Font(bold=True, size=12, color=C_HEADER_BG)

    breakdown = tasks["Status"].value_counts().reset_index()
    breakdown.columns = ["Status", "Count"]
    breakdown["% of Total"] = (breakdown["Count"] / total * 100).round(1).astype(str) + "%"

    for r_idx, row in enumerate(dataframe_to_rows(breakdown, index=False, header=True), 8):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == 8:
                header_style(cell)
            else:
                status_val = ws.cell(row=r_idx, column=1).value
                fill = STATUS_FILLS.get(status_val) if c_idx == 1 else None
                data_style(cell, r_idx, fill)

    set_col_widths(ws, {"A": 20, "B": 10, "C": 14, "D": 14, "E": 14, "F": 14})


# ── Sheet 2: Tasks ────────────────────────────────────────────────────────────

def build_tasks(wb: Workbook, tasks: pd.DataFrame) -> None:
    ws = wb.create_sheet("Tasks")
    ws.sheet_view.showGridLines = False

    col_widths = {"A": 10, "B": 35, "C": 18, "D": 14, "E": 12, "F": 14}
    write_df_to_sheet(ws, tasks, col_widths)

    # Apply status-based cell colouring to every data row
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        status_val = row[3].value  # column D = Status
        fill = STATUS_FILLS.get(status_val)
        if fill:
            for cell in row:
                cell.fill = fill

    # Completion % as percentage format
    for cell in ws["F"][1:]:
        cell.number_format = "0%"
        if cell.value is not None:
            cell.value = cell.value / 100


# ── Sheet 3: Budget ───────────────────────────────────────────────────────────

def build_budget(wb: Workbook, budget: pd.DataFrame) -> None:
    ws = wb.create_sheet("Budget")
    ws.sheet_view.showGridLines = False

    col_widths = {"A": 18, "B": 16, "C": 14, "D": 14}
    write_df_to_sheet(ws, budget, col_widths)

    # Currency format
    for col in ["B", "C", "D"]:
        for cell in ws[col][1:]:
            cell.number_format = '$#,##0.00'

    # Bar chart: Allocated vs Spent
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = "Budget: Allocated vs Spent"
    chart.y_axis.title = "Amount ($)"
    chart.x_axis.title = "Category"
    chart.style = 10
    chart.width = 20
    chart.height = 12

    n = len(budget) + 1
    data_ref = Reference(ws, min_col=2, max_col=3, min_row=1, max_row=n)
    cats_ref = Reference(ws, min_col=1, min_row=2, max_row=n)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)

    ws.add_chart(chart, "F2")


# ── Sheet 4: Team ─────────────────────────────────────────────────────────────

def build_team(wb: Workbook, team: pd.DataFrame) -> None:
    ws = wb.create_sheet("Team")
    ws.sheet_view.showGridLines = False

    col_widths = {"A": 18, "B": 14, "C": 14, "D": 14, "E": 12}
    write_df_to_sheet(ws, team, col_widths)

    # Stacked bar chart
    chart = BarChart()
    chart.type = "bar"
    chart.grouping = "stacked"
    chart.title = "Team Workload Distribution"
    chart.x_axis.title = "Tasks"
    chart.y_axis.title = "Team Member"
    chart.style = 10
    chart.width = 20
    chart.height = 12

    n = len(team) + 1
    data_ref = Reference(ws, min_col=3, max_col=5, min_row=1, max_row=n)
    cats_ref = Reference(ws, min_col=1, min_row=2, max_row=n)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)

    ws.add_chart(chart, "G2")


# ── Main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    t0 = time.perf_counter()
    print("Weekly Report Generator")
    print("-" * 40)

    tasks  = pd.read_csv(DATA / "tasks.csv")
    budget = pd.read_csv(DATA / "budget.csv")
    team   = pd.read_csv(DATA / "team.csv")
    print(f"  Loaded  {len(tasks)} tasks | {len(budget)} budget rows | {len(team)} team members")

    wb = Workbook()
    wb.remove(wb.active)  # remove default blank sheet

    build_summary(wb, tasks, budget, team)
    build_tasks(wb, tasks)
    build_budget(wb, budget)
    build_team(wb, team)

    OUTPUT.mkdir(exist_ok=True)
    out_path = OUTPUT / f"weekly_report_{date.today()}.xlsx"
    wb.save(out_path)

    elapsed = time.perf_counter() - t0
    print(f"  Report  saved -> {out_path.name}")
    print(f"  Time    {elapsed:.2f}s  (vs ~90 min manual)")
    print("-" * 40)
    print("Done.")


if __name__ == "__main__":
    main()
